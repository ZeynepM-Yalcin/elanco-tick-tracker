"""
converts excel into json so the main.py can load it into sqlite on startup
only need to run this once to output seed_data.json
"""

import json
import openpyxl

#open the spreadsheet
wb = openpyxl.load_workbook("Tick_Sightings.xlsx")
ws = wb.active

#grab the header row to use them as keys
headers = [cell.value for cell in ws[1]]

#loop through every row after the header and build a list of dicts
records = []
for row in ws.iter_rows(min_row=2, values_only=True):
    record = {}
    for i, value in enumerate(row):
        if value is not None:
            record[headers[i]] = str(value)  #convert everything to strings to keep it simple
    
    #only include rows that actually have data
    if record.get("id") and record.get("location"):
        records.append(record)

#write it all out to JSON
with open("seed_data.json", "w") as f:
    json.dump(records, f, indent=2)

print(f"Done — {len(records)} records written to seed_data.json")